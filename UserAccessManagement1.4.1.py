# -*- coding: utf-8 -*-
import sys
import os
import re
import subprocess
import pyodbc
import csv
try:
    from openpyxl import Workbook
except Exception:
    Workbook = None
from PyQt5.QtWidgets import (
    QApplication,
    QWidget,
    QPushButton,
    QVBoxLayout,
    QHBoxLayout,
    QTreeWidget,
    QTreeWidgetItem,
    QTableWidget,
    QTableWidgetItem,
    QMessageBox,
    QCheckBox,
    QLabel,
    QInputDialog,
    QFileDialog,
    QLineEdit,
    QFormLayout,
    QDesktopWidget,
    QDialog,
    QListWidget,
    QListWidgetItem,
    QDialogButtonBox,
    QHeaderView,
)
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtGui import QIcon, QBrush, QColor
from PyQt5.QtCore import QSettings
import logging
from typing import List, Optional


logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
if not logger.handlers:
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    fmt = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    ch.setFormatter(fmt)
    logger.addHandler(ch)

DEFAULT_PASSWORD = "xx17737xx"
SETTINGS_ORG = "AccessApp"
SETTINGS_APP = "UserAccessManager"

Base_Dir = getattr(sys,'_MEIPASS',os.path.dirname(os.path.abspath(__file__)))
WINDOW_ICON_PATH = os.path.join(Base_Dir,"logo.png")


def center_window(widget: QWidget) -> None:
    try:
        qr = widget.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        widget.move(qr.topLeft())
    except Exception:
        pass

def get_saved_icon_path() -> str:
    settings = QSettings(SETTINGS_ORG, SETTINGS_APP)
    return settings.value("windowIconPath", type=str)

def set_saved_icon_path(path: str) -> None:
    settings = QSettings(SETTINGS_ORG, SETTINGS_APP)
    settings.setValue("windowIconPath", path)

def apply_window_icon(widget: QWidget, icon_path: str = None) -> None:
    chosen_path = icon_path or WINDOW_ICON_PATH or get_saved_icon_path()
    if chosen_path and os.path.exists(chosen_path):
        widget.setWindowIcon(QIcon(chosen_path))

def apply_app_icon(app: QApplication) -> None:
    chosen_path = WINDOW_ICON_PATH or get_saved_icon_path()
    if chosen_path and os.path.exists(chosen_path):
        app.setWindowIcon(QIcon(chosen_path))

light_theme = """
QWidget {
    background: #f4f4f4;
    color: #1e293b;
    font-size: 13px;
}
QPushButton {
    background: #ffffff;
    color: #312e81;
    border: none;
    padding: 12px;
    border-radius: 12px;
    font-weight: bold;
}
QPushButton:hover { background: #eef2ff; }
QPushButton:disabled { background: #e2e8f0; color: #94a3b8; }
QLineEdit, QInputDialog, QTableWidget, QTableView, QTreeWidget {
    background: #ffffff;
    color: #1e293b;
    border: 1px solid #cbd5e1;
    border-radius: 8px;
    padding: 6px;
}
QHeaderView::section {
    background: #f1f5f9;
    color: #475569;
    padding: 6px;
    border: none;
    font-weight: bold;
}
QCheckBox { spacing: 8px; color: #334155; }
QMessageBox { background: #ffffff; }
QLabel { color: #1e293b; }
QScrollBar:vertical { background: #f1f5f9; width: 10px; border-radius: 4px; }
QScrollBar::handle:vertical { background: #cbd5e1; border-radius: 4px; }
QScrollBar::handle:vertical:hover { background: #94a3b8; }
"""

dark_theme = """
QWidget {
    background: #1e293b;
    color: #e2e8f0;
    font-size: 13px;
}
QPushButton {
    background: #334155;
    color: #c7d2fe;
    border: none;
    padding: 12px;
    border-radius: 12px;
    font-weight: bold;
}
QPushButton:hover {
    background: #475569;
}
QPushButton:pressed {
    background: #6366f1;
    color: #ffffff;
}
QPushButton:disabled {
    background: #1e293b;
    color: #64748b;
}
QLineEdit, QInputDialog, QTableWidget, QTableView, QTreeWidget {
    background: #0f172a;
    color: #e2e8f0;
    border: 1px solid #334155;
    border-radius: 8px;
    padding: 6px;
}
QHeaderView::section {
    background: #334155;
    color: #cbd5e1;
    padding: 6px;
    border: none;
    font-weight: bold;
}
QCheckBox {
    spacing: 8px;
    color: #e2e8f0;
}
QMessageBox {
    background: #1e293b;
    color: #f1f5f9;
}
QLabel {
    color: #e2e8f0;
}
QScrollBar:vertical {
    background: #0f172a;
    width: 10px;
    margin: 0px;
    border-radius: 4px;
}
QScrollBar::handle:vertical {
    background: #475569;
    border-radius: 4px;
}
QScrollBar::handle:vertical:hover {
    background: #6366f1;
}
"""


def normalize_persian_text(text: str) -> str:
    if not isinstance(text, str):
        return text
    replacements = {
        "\u064A": "\u06CC",
        "\u0643": "\u06A9",
        "\u0629": "\u0647",
        "\u064B": "",
        "\u064C": "",
        "\u064D": "",
        "\u064E": "",
        "\u064F": "",
        "\u0650": "",
        "\u0651": "",
        "\u0652": "",
        "\u0670": "",
        "\u0622": "\u0627",
        "\u0623": "\u0627",
        "\u0625": "\u0627",
        "\u0624": "\u0648",
        "\u06C0": "\u0647",
        "\u0640": "",
        "\u200C": "",
        "\u200F": "",
        "\u200E": "",
    }
    normalized = text
    for src, dst in replacements.items():
        normalized = normalized.replace(src, dst)
    normalized = " ".join(normalized.split())
    return normalized

def build_like_param(text: str) -> str:
    return f"%{normalize_persian_text(text)}%"

def candidate_collations() -> list:
    return [
        "Persian_100_CI_AI",
        "Arabic_100_CI_AI",
        "SQL_Latin1_General_CP1256_CI_AI",
    ]

def sql_normalize_expr(col_expr: str) -> str:
    return (
        "REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE(REPLACE("
        f"{col_expr}, N'ي', N'ی'), N'ى', N'ی'), N'ك', N'ک'), N'ۀ', N'ه'), N'ة', N'ه'), N'ـ', N''), "
        "NCHAR(8204), N''), NCHAR(8205), N'')"
    )

SERVER = r".\Moein"
DATABASE = "Moein"
USERNAME = "Sa"
PASSWORD = "arta0@"


def find_sql_instances():
    try:
        cmd = (
            'powershell "Get-ChildItem '
            "'HKLM:\\SOFTWARE\\Microsoft\\Microsoft SQL Server\\Instance Names\\SQL' "
            '| ForEach-Object { $_.Name.Split(\'\\\\\')[-1] }"'
        )
        output = subprocess.check_output(cmd, shell=True, text=True)
        instances = [inst.strip() for inst in output.splitlines() if inst.strip()]
        return instances
    except Exception:
        return []

def find_latest_moein_db(server_name, username="sa", password="arta0@"):
    try:
        conn_str = f"DRIVER={{SQL Server}};SERVER={server_name};UID={username};PWD={password}"
        conn = pyodbc.connect(conn_str, timeout=3)
        cursor = conn.cursor()
        cursor.execute(
            "SELECT name FROM sys.databases WHERE name LIKE 'Moein%' ORDER BY name ASC;"
        )
        dbs = [r[0] for r in cursor.fetchall()]
        if not dbs:
            return None
        dbs_sorted = sorted(
            dbs,
            key=lambda x: int(re.findall(r"\d+", x)[0]) if re.findall(r"\d+", x) else 0,
        )
        return dbs_sorted[-1]
    except Exception:
        return None

def auto_connect():
    possible_instances = [
        r".\Moein",
        r".\Moein2008",
        r".\Moein2012",
        r".\Moein2014",
        r".\Moein2019",
        r".\Moein2022",
    ]
    possible_instances.extend(find_sql_instances())

    for instance in possible_instances:
        db = find_latest_moein_db(instance)
        if db:
            try:
                conn_str = (
                    f"DRIVER={{SQL Server}};SERVER={instance};DATABASE={db};UID=sa;PWD=arta0@"
                )
                conn = pyodbc.connect(conn_str, autocommit=False)
                return conn, instance, db
            except Exception:
                continue
    return None, None, None

def ensure_isactive_column(conn) -> None:
    cur = conn.cursor()
    try:
        cur.execute(
            """
            SELECT 1
            FROM INFORMATION_SCHEMA.TABLES
            WHERE TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'UserAccess'
            """
        )
        tbl_exists = cur.fetchone() is not None
        if not tbl_exists:
            logger.warning("Table dbo.UserAccess not found; skip ensuring IsActive column.")
            return

        cur.execute(
            """
            SELECT 1
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'UserAccess' AND COLUMN_NAME = 'IsActive'
            """
        )
        col_exists = cur.fetchone() is not None

        if not col_exists:
            cur.execute(
                """
                ALTER TABLE dbo.UserAccess
                ADD IsActive BIT NOT NULL CONSTRAINT DF_UserAccess_IsActive DEFAULT (1) WITH VALUES;
                """
            )
            conn.commit()
            logger.info("IsActive column added to dbo.UserAccess with default 1.")
        else:
            cur.execute("UPDATE dbo.UserAccess SET IsActive = 1 WHERE IsActive IS NULL;")
            conn.commit()
            logger.info("IsActive nulls backfilled to 1 on dbo.UserAccess.")
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        logger.exception("Failed to ensure IsActive column on dbo.UserAccess")
        raise
    finally:
        try:
            cur.close()
        except Exception:
            pass

def stored_procedure_exists(conn, schema_name: str, procedure_name: str) -> bool:
    cursor = conn.cursor()
    try:
        cursor.execute(
            """
            SELECT 1
            FROM sys.procedures
            WHERE name = ? AND schema_id = SCHEMA_ID(?)
            """,
            (procedure_name, schema_name),
        )
        return cursor.fetchone() is not None
    except Exception:
        return False
    finally:
        try:
            cursor.close()
        except Exception:
            pass

def _fallback_set_user_access_single(conn, user_id: int, formbutton_id: int, is_active: bool) -> None:
    cursor = conn.cursor()
    try:
        if is_active:
            cursor.execute(
                """
                IF NOT EXISTS (SELECT 1 FROM dbo.UserAccess WHERE UserId = ? AND FormButtonsId = ?)
                INSERT INTO dbo.UserAccess (UserId, FormButtonsId, IsActive) VALUES (?, ?, 1);
                ELSE
                UPDATE dbo.UserAccess SET IsActive = 1 WHERE UserId = ? AND FormButtonsId = ?;
                """,
                (user_id, formbutton_id, user_id, formbutton_id, user_id, formbutton_id),
            )
        else:
            cursor.execute(
                """
                DELETE FROM dbo.UserAccess
                WHERE UserId = ? AND FormButtonsId = ?;
                """,
                (user_id, formbutton_id),
            )
        conn.commit()
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        raise
    finally:
        try:
            cursor.close()
        except Exception:
            pass

def _fallback_set_form_access_for_user(conn, user_id: int, form_id: int, is_active: bool) -> None:
    cursor = conn.cursor()
    try:
        if is_active:
            cursor.execute(
                """
                UPDATE ua
                SET ua.IsActive = 1
                FROM dbo.UserAccess AS ua
                JOIN dbo.FormButtons AS fb ON fb.ID = ua.FormButtonsId
                WHERE ua.UserId = ? AND fb.IDForm = ?;
                """,
                (user_id, form_id),
            )

            cursor.execute(
                """
                INSERT INTO dbo.UserAccess (UserId, FormButtonsId, IsActive)
                SELECT ?, fb.ID, 1
                FROM dbo.FormButtons AS fb
                WHERE fb.IDForm = ?
                  AND NOT EXISTS (
                      SELECT 1 FROM dbo.UserAccess AS ua
                      WHERE ua.UserId = ? AND ua.FormButtonsId = fb.ID
                  );
                """,
                (user_id, form_id, user_id),
            )
        else:
            cursor.execute(
                """
                DELETE ua
                FROM dbo.UserAccess AS ua
                JOIN dbo.FormButtons AS fb ON fb.ID = ua.FormButtonsId
                WHERE ua.UserId = ? AND fb.IDForm = ?;
                """,
                (user_id, form_id),
            )
        conn.commit()
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        raise
    finally:
        try:
            cursor.close()
        except Exception:
            pass

def _fallback_set_user_access_rewrite(conn, user_id: int, formbutton_ids: List[int]) -> None:
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM dbo.UserAccess WHERE UserId = ?;", (user_id,))
        ids_to_insert = [(user_id, int(fid)) for fid in (formbutton_ids or [])]
        if ids_to_insert:
            cursor.executemany(
                "INSERT INTO dbo.UserAccess (UserId, FormButtonsId, IsActive) VALUES (?, ?, 1);",
                ids_to_insert,
            )
        conn.commit()
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        raise
    finally:
        try:
            cursor.close()
        except Exception:
            pass

def _exec_proc(conn, proc_call: str, params: List):
    cur = conn.cursor()
    try:
        cur.execute(proc_call, params)
        conn.commit()
        return True, None
    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        logger.exception('Error executing proc: %s', proc_call)
        return False, str(e)
    finally:
        try:
            cur.close()
        except Exception:
            pass

def set_user_access_rewrite(conn, user_id: int, formbutton_ids: List[int], changed_by: Optional[int] = None):
    ids_str = ','.join(str(int(x)) for x in formbutton_ids) if formbutton_ids else ''
    if stored_procedure_exists(conn, 'dbo', 'SetUserAccess_Rewrite'):
        proc = 'EXEC dbo.SetUserAccess_Rewrite ?, ?, ?'
        params = (user_id, ids_str, changed_by)
        ok, err = _exec_proc(conn, proc, params)
        if not ok:
            raise RuntimeError(f'SetUserAccess_Rewrite failed: {err}')
    else:
        _fallback_set_user_access_rewrite(conn, user_id, formbutton_ids)
    logger.info('SetUserAccess_Rewrite succeeded for UserId=%s (count=%d)', user_id, len(formbutton_ids))

def set_user_access_single(conn, user_id: int, formbutton_id: int, is_active: bool, changed_by: Optional[int] = None):
    _fallback_set_user_access_single(conn, user_id, formbutton_id, is_active)
    logger.info('SetUserAccess_Single done: UserId=%s FB=%s Active=%s', user_id, formbutton_id, is_active)

def set_form_access_for_user(conn, user_id: int, form_id: int, is_active: bool, changed_by: Optional[int] = None):
    _fallback_set_form_access_for_user(conn, user_id, form_id, is_active)
    logger.info('SetFormAccess_ForUser done: UserId=%s FormId=%s Active=%s', user_id, form_id, is_active)

class LoginWindow(QWidget):
    def __init__(self, icon_path: str = None):
        super().__init__()
        self.setWindowTitle("ورود به برنامه")
        self.setGeometry(500, 300, 380, 160)
        self.setLayoutDirection(Qt.RightToLeft)
        apply_window_icon(self, icon_path)
        center_window(self)

        layout = QVBoxLayout()
        form = QFormLayout()
        self.txt_password = QLineEdit()
        self.txt_password.setEchoMode(QLineEdit.Password)
        self.txt_password.setPlaceholderText("رمز عبور را وارد کنید")
        form.addRow("رمز عبور:", self.txt_password)
        layout.addLayout(form)

        self.btn_login = QPushButton("ورود")
        self.btn_login.clicked.connect(self.handle_login)
        layout.addWidget(self.btn_login, alignment=Qt.AlignCenter)

        self.setLayout(layout)

    def handle_login(self):
        entered = self.txt_password.text().strip()
        if entered == DEFAULT_PASSWORD:
            self.close()
            self.auto = AutoConnectWindow()
            self.auto.show()
        else:
            QMessageBox.warning(self, "رمز نادرست", "رمز عبور صحیح نیست.")

class ManualConnectWindow(QWidget):
    def __init__(self, parent=None, icon_path: str = None):
        super().__init__(parent)
        self.setWindowTitle("تنظیمات اتصال به SQL Server (اتصال دستی)")
        self.setGeometry(500, 250, 400, 260)
        self.setLayoutDirection(Qt.RightToLeft)
        apply_window_icon(self, icon_path)
        center_window(self)

        layout = QVBoxLayout()

        form = QFormLayout()
        self.txt_server = QLineEdit(r".\Moein")
        self.txt_db = QLineEdit("Moein")
        self.txt_user = QLineEdit("sa")
        self.txt_pass = QLineEdit("arta0@")
        self.txt_pass.setEchoMode(QLineEdit.Password)

        form.addRow("Server Name:", self.txt_server)
        form.addRow("Database Name:", self.txt_db)
        form.addRow("Username:", self.txt_user)
        form.addRow("Password:", self.txt_pass)

        layout.addLayout(form)

        self.btn_connect = QPushButton("اتصال به دیتابیس")
        self.btn_connect.clicked.connect(self.connect_to_db)
        layout.addWidget(self.btn_connect, alignment=Qt.AlignCenter)

        self.setLayout(layout)

    def connect_to_db(self):
        server = self.txt_server.text().strip()
        db = self.txt_db.text().strip()
        user = self.txt_user.text().strip()
        pwd = self.txt_pass.text().strip()

        try:
            conn_str = f"DRIVER={{SQL Server}};SERVER={server};DATABASE={db};UID={user};PWD={pwd}"
            conn = pyodbc.connect(conn_str, autocommit=False)
            ensure_isactive_column(conn)

            QMessageBox.information(self, "موفقیت ✅", f"اتصال موفق به دیتابیس '{db}' برقرار شد.")
            self.close()
            self.main_window = MainWindow(conn)
            self.main_window.show()

        except Exception as e:
            QMessageBox.critical(self, "خطا ❌", f"اتصال ناموفق بود:\n{str(e)}")

class AutoConnectWindow(QWidget):
    def __init__(self, icon_path: str = None):
        super().__init__()
        self.setWindowTitle("اتصال خودکار به SQL Server")
        self.setGeometry(500, 300, 380, 120)
        self.setLayoutDirection(Qt.RightToLeft)
        apply_window_icon(self, icon_path)
        center_window(self)

        layout = QVBoxLayout()
        self.label = QLabel("در حال تلاش برای اتصال خودکار به دیتابیس Moein ...")
        self.label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.label)
        self.setLayout(layout)

        QTimer.singleShot(100, self.try_auto_connect)

    def try_auto_connect(self):
        conn, server, db = auto_connect()
        if conn:
            try:
                ensure_isactive_column(conn)
            except Exception as e:
                QMessageBox.warning(self, "خطای سازگارسازی اسکیم", f"عدم موفقیت در تضمین ستون IsActive:\n{str(e)}")
            QMessageBox.information(self, "موفقیت ✅", f"اتصال خودکار برقرار شد:\n{server} → {db}")
            self.close()
            self.main_window = MainWindow(conn)
            self.main_window.show()
        else:
            QMessageBox.warning(
                self,
                "اتصال خودکار ناموفق ⚠️",
                "هیچ دیتابیس Moein پیدا نشد. لطفاً اتصال دستی را انجام دهید.",
            )
            self.close()
            self.manual = ManualConnectWindow()
            self.manual.show()

class UserSelectDialog(QDialog):
    def __init__(self, main_window: 'MainWindow'):
        super().__init__(main_window)
        self.main = main_window
        self._selected = None

        self.setWindowTitle("انتخاب کاربر")
        self.setGeometry(480, 260, 440, 520)
        self.setLayoutDirection(Qt.RightToLeft)
        apply_window_icon(self)
        center_window(self)

        layout = QVBoxLayout()

        self.search_box = QLineEdit()
        self.search_box.setPlaceholderText("جست‌وجوی کاربر (اختیاری)")
        layout.addWidget(self.search_box)

        self.list_widget = QListWidget()
        self.list_widget.setLayoutDirection(Qt.RightToLeft)
        layout.addWidget(self.list_widget)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self._handle_accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

        self.setLayout(layout)

        self.search_box.textChanged.connect(self._on_search_text_changed)
        self.list_widget.itemDoubleClicked.connect(self._handle_item_double_clicked)

        try:
            users = self.main.query_users_initial()
            self._populate(users)
        except Exception as e:
            QMessageBox.critical(self, "خطا", f"خطا در بارگذاری کاربران:\n{str(e)}")

    def _populate(self, users: list):
        self.list_widget.clear()
        for uid, uname in users:
            text = f"{uid} - {uname}"
            item = QListWidgetItem(text)
            item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            item.setData(Qt.UserRole, (uid, uname))
            self.list_widget.addItem(item)

    def _on_search_text_changed(self, text: str):
        try:
            t = (text or "").strip()
            if t:
                users = self.main.query_users_by_name(t)
            else:
                users = self.main.query_users_initial()
            self._populate(users)
        except Exception as e:
            QMessageBox.critical(self, "خطا", f"خطا در جست‌وجوی کاربر:\n{str(e)}")

    def _handle_accept(self):
        item = self.list_widget.currentItem()
        if not item:
            QMessageBox.warning(self, "انتخاب کاربر", "یک کاربر را انتخاب کنید.")
            return
        uid, uname = item.data(Qt.UserRole)
        self._selected = (uid, uname)
        self.accept()

    def _handle_item_double_clicked(self, item: QListWidgetItem):
        if item is None:
            return
        uid, uname = item.data(Qt.UserRole)
        self._selected = (uid, uname)
        self.accept()

    def selected_user(self):
        return self._selected if self._selected is not None else (None, "")

MENU_GROUPS = {
    "معرفی اطلاعات": [
        "معرفی اشخاص",
        "معرفی کالا",
        "معرفی خدمات",
        "معرفی بانک",
        "معرفی دسته چک ها",
        "چک های دریافتی اول دوره",
        "چک های پرداختی اول دوره",
        "معرفی صندوق ها",
        "معرفی انبار",
        "معرفی هزینه ها",
        "معرفی درامد",
        "لیست سطوح قیمتی"
    ],
    "عملیات": [
        "صدور فیش",
        "پیش فاکتور",
        "فاکتور خرید",
        "فاکتور فروش",
        "فاکتور خدمات",
        "فاکتور برگشت از خرید",
        "فاکتور برگشت از فروش",
        "عملیات دریافت",
        "عملیات پرداخت",
        "انتقال وجه",
        "ثبت هزینه",
        "ثبت درامد",
        "چک های دریافتی",
        "چک های پرداختی",
        "حواله بین انبار",
        "اقساط دریافتی"
    ],
    "گزارشات": [
        "گزارش اشخاص",
        "گزارش بانک ها",
        "گزارش صندوق ها",
        "موجودی کالا",
        "گزارش فروش",
        "گزارش سود و زیان فاکتور ها",
        "گزارش سود و زیان",
        "ترازنامه",
        "یاداوری",
        "گزارش کابران",
        "گزراش ویزیتور",
        "گزارش فصلی",
        "دفتر روزانه"
    ],
    "حسابدری": [
        "سرفصل های حسابداری",
        "کدینگ حسابدرای",
        "صورت حساب سود و زیان",
        "ترازنامه",
        "تراز2ستونی",
        "تراز4ستونی",
        "دفترکل",
        "گزارش معین",
        "گزارش تفضیلی",
        "گزارش اسناد",
        "دفترروزنامه",
        "مانده حساب های اشخاص بر اساس اسناد",
        "تایید سند",
        "عدم تایید سند",
        "جستجوی اسناد"
    ],
    "تولید": [
        "فرمول تولید",
        "تولید کالا"
    ],
    "حقوق و دستمزد": [
        "معرفی کارمندان",
        "معرفی کارمندان"
    ],
    "ابزار": [
        "پشتیبانی از راه دور",
        "پشتیبان گیری اطلاعات",
        "بازگردانی فایل پشتیبان",
        "پیامک رسان",
        "دفترچه تلفن",
        "ماشین حساب",
        "چاپ بارکد",
        "چاپ چک",
        "تنظیمات چاپ",
        "دفترپیگری"
    ],
    "تنظیمات": [
        "معرفی شرکت",
        "معرفی گروه های دسترسی",
        "معرفی کاربران",
        "میزکار",
        "تغییر ظاهر برنامه",
        "تنظیمات بانک اطلاعاتی",
        "تنظیمات فاکتورها",
        "تنظیمات یاداوری ها"
    ],
    "راهنما": [
        "راهنما",
        "اطلاعات نرم افزار",
        "درباره ما"
    ],
    "خروج": [
        "خروج موقت",
        "تغییر کاربر"
    ]
}

class MainWindow(QWidget):
    def __init__(self, connection, icon_path: str = None):
        super().__init__()
        self.conn = connection
        self.current_user_id = None
        self.current_user_name = ""
        self.current_mode = None
        self._last_visible_forms = {}
        self._last_user_id = None
        self._last_is_allowed = None
        self._button_id_to_item = {}
        self._all_forms_dict = {}
        self.is_dark = True

        self.setWindowTitle("هامدیریت دسترسی")
        self.setGeometry(250, 80, 1200, 700)
        self.setLayoutDirection(Qt.RightToLeft)
        apply_window_icon(self, icon_path)
        center_window(self)

        layout = QVBoxLayout()

        self.search_box = QLineEdit()
        self.search_box.setPlaceholderText("جست‌وجو (کلمه به کلمه)")
        self.search_box.textChanged.connect(self.filter_tree_and_table)
        layout.addWidget(self.search_box)

        button_layout = QVBoxLayout()
        self.btn_show_allowed = QPushButton("نمایش دسترسی های غیر مجاز کاربر")
        self.btn_show_denied = QPushButton("نمایش دسترسی های مجاز کاربر")
        self.btn_show_all = QPushButton("نمایش همه فرم‌ها")
        self.btn_save_all = QPushButton("ذخیره تغییرات")
        self.btn_export_excel = QPushButton("خروجی اکسل/CSV")
        self.btn_toggle_theme = QPushButton("تغییر تم")
        self.btn_toggle_theme.clicked.connect(self.toggle_theme)
        button_layout.addWidget(self.btn_show_allowed)
        button_layout.addWidget(self.btn_show_denied)
        button_layout.addWidget(self.btn_show_all)
        button_layout.addWidget(self.btn_save_all)
        button_layout.addWidget(self.btn_export_excel)
        button_layout.addWidget(self.btn_toggle_theme)

        main_layout = QHBoxLayout()
        
        self.tree_widget = QTreeWidget()
        self.tree_widget.setHeaderLabels(["دسترسی", "وضعیت"])
        self.tree_widget.setColumnWidth(0, 600)
        # self.tree_widget.setLayoutDirection(Qt.RightToLeft)
        self.tree_widget.itemChanged.connect(self.on_tree_item_changed)
        self.tree_widget.currentItemChanged.connect(self.on_tree_selection_changed)
        main_layout.addWidget(self.tree_widget, 1)

        self.table_widget = QTableWidget()
        self.table_widget.setColumnCount(2)
        self.table_widget.setHorizontalHeaderLabels(["کد", "نام"])
        # self.table_widget.setLayoutDirection(Qt.RightToLeft)
        self.table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        main_layout.addWidget(self.table_widget, 1)

        layout.addLayout(button_layout)
        layout.addLayout(main_layout)

        self.lbl_status = QLabel("")
        layout.addWidget(self.lbl_status)

        self.setLayout(layout)

        self.btn_show_allowed.clicked.connect(self.show_allowed_forms)
        self.btn_show_denied.clicked.connect(self.show_denied_forms)
        self.btn_show_all.clicked.connect(self.show_all_forms)
        self.btn_save_all.clicked.connect(self.save_all_changes)
        self.btn_export_excel.clicked.connect(self.export_current_view_to_file)
        self.update_export_button_state()

        QTimer.singleShot(0, self.select_user_workflow)

    def toggle_theme(self):
        self.is_dark = not self.is_dark
        app = QApplication.instance()
        app.setStyleSheet(dark_theme if self.is_dark else light_theme)

    def on_tree_item_changed(self, item, column):
        if column != 0:
            return
        
        item_type = item.data(0, Qt.UserRole + 1)
        if item_type == "button":
            button_id = item.data(0, Qt.UserRole)
            is_checked = item.checkState(0) == Qt.Checked
            parent = item.parent()
            if parent:
                self.update_parent_state(parent)
                grandparent = parent.parent()
                if grandparent:
                    self.update_parent_state(grandparent)
        elif item_type == "form":
            form_id = item.data(0, Qt.UserRole)
            is_checked = item.checkState(0) == Qt.Checked
            self.tree_widget.blockSignals(True)
            for i in range(item.childCount()):
                child = item.child(i)
                self._set_children_check_state_recursive(child, Qt.Checked if is_checked else Qt.Unchecked)
            self.tree_widget.blockSignals(False)
            parent = item.parent()
            if parent:
                self.update_parent_state(parent)
        elif item_type == "group":
            is_checked = item.checkState(0) == Qt.Checked
            self.tree_widget.blockSignals(True)
            for i in range(item.childCount()):
                child = item.child(i)
                if child.flags() & Qt.ItemIsUserCheckable:
                    child.setCheckState(0, Qt.Checked if is_checked else Qt.Unchecked)
                if child.data(0, Qt.UserRole + 1) == "form":
                    for k in range(child.childCount()):
                        grandchild = child.child(k)
                        grandchild.setCheckState(0, Qt.Checked if is_checked else Qt.Unchecked)
            self.tree_widget.blockSignals(False)

    def update_parent_state(self, parent_item):
        if not parent_item:
            return
        
        checked_count = 0
        partially_checked = False
        total_checkable_children = 0
        
        for i in range(parent_item.childCount()):
            child = parent_item.child(i)
            if child.flags() & Qt.ItemIsUserCheckable:
                total_checkable_children += 1
                state = child.checkState(0)
                if state == Qt.Checked:
                    checked_count += 1
                elif state == Qt.PartiallyChecked:
                    partially_checked = True

        if total_checkable_children == 0:
            parent_item.setFlags(parent_item.flags() & ~Qt.ItemIsUserCheckable)
            parent_item.setCheckState(0, Qt.Unchecked)
            return
        else:
            if not (parent_item.flags() & Qt.ItemIsUserCheckable):
                parent_item.setFlags(parent_item.flags() | Qt.ItemIsUserCheckable)
        
        self.tree_widget.blockSignals(True)
        if partially_checked:
            parent_item.setCheckState(0, Qt.PartiallyChecked)
        elif checked_count == 0:
            parent_item.setCheckState(0, Qt.Unchecked)
        elif checked_count == total_checkable_children:
            parent_item.setCheckState(0, Qt.Checked)
        else:
            parent_item.setCheckState(0, Qt.PartiallyChecked)
        self.tree_widget.blockSignals(False)

    def _set_children_check_state_recursive(self, item, state):
        if item is None:
            return
        if item.flags() & Qt.ItemIsUserCheckable:
            item.setCheckState(0, state)
        for i in range(item.childCount()):
            self._set_children_check_state_recursive(item.child(i), state)

    def on_tree_selection_changed(self, current, previous):
        try:
            if current is None:
                return
            item_type = current.data(0, Qt.UserRole + 1)
            if item_type == "form":
                form_id = current.data(0, Qt.UserRole)
                form_name = current.text(0)
                buttons = []
                form_info = self._last_visible_forms.get(form_id)
                if form_info:
                    buttons = form_info.get('buttons', [])
                self.table_widget.setRowCount(len(buttons))
                self.table_widget.setColumnCount(2)
                self.table_widget.setHorizontalHeaderLabels(["کد", "نام"])
                row = 0
                for btn in buttons:
                    self.table_widget.setItem(row, 0, QTableWidgetItem(str(btn['id'])))
                    self.table_widget.setItem(row, 1, QTableWidgetItem(f"{form_name} - {btn['name']}"))
                    row += 1
            elif item_type == "button":
                button_id = current.data(0, Qt.UserRole)
                form_item = current.parent()
                form_name = form_item.text(0) if form_item else ""
                button_name = current.text(0)
                self.table_widget.setRowCount(1)
                self.table_widget.setColumnCount(2)
                self.table_widget.setHorizontalHeaderLabels(["کد", "نام"])
                self.table_widget.setItem(0, 0, QTableWidgetItem(str(button_id)))
                self.table_widget.setItem(0, 1, QTableWidgetItem(f"{form_name} - {button_name}"))
            elif item_type == "group":
                self.table_widget.setRowCount(0)
        except Exception:
            pass

    def update_export_button_state(self):
        self.btn_export_excel.setEnabled(self.current_mode in ("allowed", "denied"))

    def save_all_changes(self):
        if self.current_user_id is None:
            QMessageBox.warning(self, "انتخاب کاربر", "ابتدا یک کاربر انتخاب کنید.")
            return

        checked_button_ids = []
        if self._all_forms_dict:
            for form_id, form_data in self._all_forms_dict.items():
                for btn in form_data.get('buttons', []):
                    btn_id = btn.get('id')
                    tree_item = self._button_id_to_item.get(btn_id)
                    if tree_item is not None:
                        is_checked = tree_item.checkState(0) == Qt.Checked
                    else:
                        is_checked = bool(btn.get('access'))
                    if is_checked:
                        checked_button_ids.append(btn_id)
        else:
            for i in range(self.tree_widget.topLevelItemCount()):
                group_item = self.tree_widget.topLevelItem(i)
                for j in range(group_item.childCount()):
                    form_item = group_item.child(j)
                    for k in range(form_item.childCount()):
                        button_item = form_item.child(k)
                        if button_item.flags() & Qt.ItemIsUserCheckable and button_item.checkState(0) == Qt.Checked:
                            button_id = button_item.data(0, Qt.UserRole)
                            if button_id:
                                checked_button_ids.append(button_id)

        try:
            set_user_access_rewrite(self.conn, self.current_user_id, checked_button_ids)
            self.notify_saved("همه تغییرات ذخیره شد")
            self.refresh_main_app()
            QTimer.singleShot(100, self.reload_current_mode)
        except Exception as e:
            QMessageBox.critical(self, "خطا در ذخیره", f"خطا هنگام ذخیره تغییرات:\n{str(e)}")

    def refresh_main_app(self):
        try:
            subprocess.call(['taskkill', '/IM', 'Moein.exe', '/F'], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            main_app_path = None
            possible_paths = [
                r"C:\Program Files (x86)\MoeinSystem\Moein.exe",
                r"C:\Program Files\MoeinSystem\Moein.exe",
                r"C:\Moein\Moein.exe",
            ]
            for path in possible_paths:
                if os.path.exists(path):
                    main_app_path = path
                    break
            if main_app_path:
                subprocess.Popen(main_app_path)
                logger.info(f"برنامه اصلی Moein.exe در مسیر {main_app_path} اجرا شد.")
            else:
                logger.warning("مسیر برنامه اصلی Moein.exe یافت نشد. لطفا مسیر را در کد تنظیم کنید.")
        except Exception as e:
            logger.exception("خطا در رفرش برنامه اصلی: %s", str(e))

    def show_allowed_forms(self):
        user_id, ok = self.ask_user_id()
        if not ok:
            return
        self.current_mode = 'allowed'
        self.load_hierarchy(user_id, True)
        self.update_export_button_state()

    def show_denied_forms(self):
        user_id, ok = self.ask_user_id()
        if not ok:
            return
        self.current_mode = 'denied'
        self.load_hierarchy(user_id, False)
        self.update_export_button_state()

    def show_all_forms(self):
        user_id, ok = self.ask_user_id()
        if not ok:
            return
        self.current_mode = 'all'
        self.load_hierarchy(user_id, None)
        self.update_export_button_state()

    def load_hierarchy(self, user_id, is_allowed):
        self.tree_widget.clear()
        self.table_widget.setRowCount(0)
        
        self.tree_widget.blockSignals(True)
        
        try:
            cursor = self.conn.cursor()
            
            query = """
            SELECT 
                f.ID as FormId,
                f.Name as FormName,
                fb.ID as ButtonId,
                fb.Name as ButtonName,
                ISNULL(ua.IsActive, 0) as HasAccess
            FROM dbo.Forms f
            LEFT JOIN dbo.FormButtons fb ON fb.IDForm = f.ID
            LEFT JOIN dbo.UserAccess ua ON ua.FormButtonsId = fb.ID AND ua.UserId = ?
            ORDER BY f.Name, fb.Name;
            """
            cursor.execute(query, (user_id,))
            rows = cursor.fetchall()

            forms_dict = {}
            for row in rows:
                form_id, form_name, button_id, button_name, has_access = row
                
                if form_id not in forms_dict:
                    forms_dict[form_id] = {
                        'name': form_name,
                        'buttons': []
                    }
                
                if button_id:
                    forms_dict[form_id]['buttons'].append({
                        'id': button_id,
                        'name': button_name,
                        'access': has_access
                    })

            self._button_id_to_item = {}
            self._all_forms_dict = forms_dict

            visible_forms = {}
            matched_forms = set()
            
            for group_name, sub_forms in MENU_GROUPS.items():
                group_item = QTreeWidgetItem(self.tree_widget, [group_name, ""])
                group_item.setData(0, Qt.UserRole + 1, "group")
                group_item.setFlags(group_item.flags() | Qt.ItemIsUserCheckable)
                group_item.setCheckState(0, Qt.Unchecked)

                group_checked_count = 0
                group_total = 0

                for sub_form_name in sub_forms:
                    form_id = next((fid for fid, fdata in forms_dict.items() 
                                  if normalize_persian_text(fdata['name']) == normalize_persian_text(sub_form_name)), None)
                    if form_id is None:
                        continue

                    matched_forms.add(form_id)
                    form_data = forms_dict[form_id]

                    if is_allowed is not None:
                        has_any_access = any(btn['access'] for btn in form_data['buttons']) if form_data['buttons'] else False
                        if is_allowed and not has_any_access and form_data['buttons']:
                            continue
                        if not is_allowed and has_any_access:
                            continue

                    form_item = QTreeWidgetItem(group_item, [form_data['name'], ""])
                    form_item.setData(0, Qt.UserRole, form_id)
                    form_item.setData(0, Qt.UserRole + 1, "form")
                    form_item.setFlags(form_item.flags() | Qt.ItemIsUserCheckable)
                    form_item.setCheckState(0, Qt.Unchecked)

                    checked_count = 0
                    has_buttons = False
                    for button in form_data['buttons']:
                        has_buttons = True
                        button_item = QTreeWidgetItem(form_item, [button['name'], ""])
                        button_item.setData(0, Qt.UserRole, button['id'])
                        button_item.setData(0, Qt.UserRole + 1, "button")
                        self._button_id_to_item[button['id']] = button_item
                        button_item.setFlags(button_item.flags() | Qt.ItemIsUserCheckable)
                        button_item.setCheckState(0, Qt.Checked if button['access'] else Qt.Unchecked)
                        if button['access']:
                            checked_count += 1
                        
                        if button['access']:
                            button_item.setForeground(0, QBrush(QColor("#10b981")))
                        else:
                            button_item.setForeground(0, QBrush(QColor("#ef4444")))

                    if has_buttons:
                        total_buttons = len(form_data['buttons'])
                        group_total += total_buttons
                        group_checked_count += checked_count
                        if checked_count == 0:
                            form_item.setCheckState(0, Qt.Unchecked)
                        elif checked_count == total_buttons:
                            form_item.setCheckState(0, Qt.Checked)
                        else:
                            form_item.setCheckState(0, Qt.PartiallyChecked)
                    else:
                        form_item.setFlags(form_item.flags() & ~Qt.ItemIsUserCheckable)
                        form_item.setCheckState(0, Qt.Unchecked)

                    visible_forms[form_id] = form_data

                if group_total > 0:
                    if group_checked_count == 0:
                        group_item.setCheckState(0, Qt.Unchecked)
                    elif group_checked_count == group_total:
                        group_item.setCheckState(0, Qt.Checked)
                    else:
                        group_item.setCheckState(0, Qt.PartiallyChecked)
                else:
                    group_item.setHidden(True)

            other_group_item = QTreeWidgetItem(self.tree_widget, ["سایر", ""])
            other_group_item.setData(0, Qt.UserRole + 1, "group")
            other_group_item.setFlags(other_group_item.flags() | Qt.ItemIsUserCheckable)
            other_group_item.setCheckState(0, Qt.Unchecked)

            other_checked_count = 0
            other_total = 0

            for form_id, form_data in sorted(forms_dict.items(), key=lambda x: x[1]['name']):
                if form_id in matched_forms:
                    continue

                if is_allowed is not None:
                    has_any_access = any(btn['access'] for btn in form_data['buttons']) if form_data['buttons'] else False
                    if is_allowed and not has_any_access and form_data['buttons']:
                        continue
                    if not is_allowed and has_any_access:
                        continue

                form_item = QTreeWidgetItem(other_group_item, [form_data['name'], ""])
                form_item.setData(0, Qt.UserRole, form_id)
                form_item.setData(0, Qt.UserRole + 1, "form")
                form_item.setFlags(form_item.flags() | Qt.ItemIsUserCheckable)
                form_item.setCheckState(0, Qt.Unchecked)

                checked_count = 0
                has_buttons = False
                for button in form_data['buttons']:
                    has_buttons = True
                    button_item = QTreeWidgetItem(form_item, [button['name'], ""])
                    button_item.setData(0, Qt.UserRole, button['id'])
                    button_item.setData(0, Qt.UserRole + 1, "button")
                    self._button_id_to_item[button['id']] = button_item
                    button_item.setFlags(button_item.flags() | Qt.ItemIsUserCheckable)
                    button_item.setCheckState(0, Qt.Checked if button['access'] else Qt.Unchecked)
                    if button['access']:
                        checked_count += 1
                    
                    if button['access']:
                        button_item.setForeground(0, QBrush(QColor("#10b981")))
                    else:
                        button_item.setForeground(0, QBrush(QColor("#ef4444")))

                if has_buttons:
                    total_buttons = len(form_data['buttons'])
                    other_total += total_buttons
                    other_checked_count += checked_count
                    if checked_count == 0:
                        form_item.setCheckState(0, Qt.Unchecked)
                    elif checked_count == total_buttons:
                        form_item.setCheckState(0, Qt.Checked)
                    else:
                        form_item.setCheckState(0, Qt.PartiallyChecked)
                else:
                    form_item.setFlags(form_item.flags() & ~Qt.ItemIsUserCheckable)
                    form_item.setCheckState(0, Qt.Unchecked)

                visible_forms[form_id] = form_data

            if other_total > 0:
                if other_checked_count == 0:
                    other_group_item.setCheckState(0, Qt.Unchecked)
                elif other_checked_count == other_total:
                    other_group_item.setCheckState(0, Qt.Checked)
                else:
                    other_group_item.setCheckState(0, Qt.PartiallyChecked)
            else:
                other_group_item.setHidden(True)

            self.tree_widget.expandAll()

            button_count = sum(len(f['buttons']) for f in visible_forms.values())
            self.table_widget.setRowCount(button_count)
            row_idx = 0
            for form_id, form_data in visible_forms.items():
                for button in form_data['buttons']:
                    self.table_widget.setItem(row_idx, 0, QTableWidgetItem(str(button['id'])))
                    self.table_widget.setItem(row_idx, 1, QTableWidgetItem(f"{form_data['name']} - {button['name']}"))
                    row_idx += 1

            self._last_visible_forms = visible_forms
            self._last_user_id = user_id
            self._last_is_allowed = is_allowed

        except Exception as e:
            QMessageBox.critical(self, "خطا در اجرا", f"خطا هنگام بارگذاری داده:\n{str(e)}")
        finally:
            self.tree_widget.blockSignals(False)

    def filter_tree_and_table(self, text: str):
        norm_text = normalize_persian_text(text or "")
        tokens = [t.strip().lower() for t in norm_text.split() if t.strip()]

        def full_path_text(item):
            names = []
            cur = item
            while cur is not None:
                names.append(cur.text(0) or "")
                cur = cur.parent()
            names.reverse()
            return normalize_persian_text(" / ".join(names)).lower()

        def recurse(item):
            if item is None:
                return False
            item_text = full_path_text(item)
            match_self = all(tok in item_text for tok in tokens)
            any_child = False
            for i in range(item.childCount()):
                if recurse(item.child(i)):
                    any_child = True
            match = match_self or any_child or (len(tokens) == 0)
            item.setHidden(not match)
            if match:
                self.tree_widget.expandItem(item)
            return match

        for i in range(self.tree_widget.topLevelItemCount()):
            recurse(self.tree_widget.topLevelItem(i))

        for i in range(self.table_widget.rowCount()):
            cell_texts = []
            for j in range(self.table_widget.columnCount()):
                cell = self.table_widget.item(i, j)
                if cell is not None:
                    cell_texts.append(cell.text())
            row_text = normalize_persian_text(" ".join(cell_texts)).lower()
            visible = all(tok in row_text for tok in tokens)
            self.table_widget.setRowHidden(i, not visible)

    def export_current_view_to_file(self):
        if self.current_mode not in ("allowed", "denied"):
            QMessageBox.warning(self, "خروجی اکسل", "لطفاً ابتدا نمای فرم‌های مجاز یا غیرمجاز را انتخاب کنید.")
            return
        filter_str = "Excel Files (*.xlsx);;CSV Files (*.csv)"
        default_name = "خروجی_دسترسی.xlsx" if Workbook else "خروجی_دسترسی.csv"
        path, selected_filter = QFileDialog.getSaveFileName(self, "ذخیره خروجی", default_name, filter_str)
        if not path:
            return
        rows = self._collect_current_view_rows()
        try:
            if path.lower().endswith('.xlsx') and Workbook:
                wb = Workbook()
                ws_users = wb.create_sheet(title="Users")
                ws_users.append(["UserId", "UserName"])
                ws_users.append([self.current_user_id, self.current_user_name])

                ws_access = wb.create_sheet(title="Access")
                ws_access.append(["FormId", "FormName", "ButtonId", "ButtonName", "Access"])
                for r in rows:
                    ws_access.append(r[2:])

                if 'Sheet' in wb.sheetnames:
                    wb.remove(wb['Sheet'])

                wb.save(path)
                QMessageBox.information(self, "خروجی اکسل", "فایل Excel با شیت‌های جدا (Users و Access) ذخیره شد.")
            else:
                if not path.lower().endswith('.csv'):
                    path = path + '.csv'
                with open(path, 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f)
                    writer.writerow(["UserId", "UserName", "FormId", "FormName", "ButtonId", "ButtonName", "Access"])
                    writer.writerows(rows)
                QMessageBox.information(self, "خروجی اکسل", "فایل CSV با موفقیت ذخیره شد (برای شیت‌های جدا، از xlsx استفاده کنید).")
        except Exception as e:
            QMessageBox.critical(self, "خروجی اکسل", f"خطا در ذخیره فایل:\n{str(e)}")

    def _collect_current_view_rows(self):
        output_rows = []
        for i in range(self.tree_widget.topLevelItemCount()):
            group_item = self.tree_widget.topLevelItem(i)
            if group_item.isHidden():
                continue
            for j in range(group_item.childCount()):
                form_item = group_item.child(j)
                if form_item.isHidden():
                    continue
                form_id = form_item.data(0, Qt.UserRole)
                form_name = form_item.text(0)
                for k in range(form_item.childCount()):
                    button_item = form_item.child(k)
                    if button_item.isHidden():
                        continue
                    button_id = button_item.data(0, Qt.UserRole)
                    button_name = button_item.text(0)
                    access = 1 if button_item.checkState(0) == Qt.Checked else 0
                    output_rows.append([self.current_user_id, self.current_user_name, form_id, form_name, button_id, button_name, access])
        return output_rows

    def reload_current_mode(self):
        if self.current_mode == 'allowed':
            self.show_allowed_forms()
        elif self.current_mode == 'denied':
            self.show_denied_forms()
        elif self.current_mode == 'all':
            self.show_all_forms()

    def notify_saved(self, message: str):
        self.lbl_status.setText(message)
        QTimer.singleShot(1200, lambda: self.lbl_status.setText(""))

    def ask_user_id(self):
        if self.current_user_id is not None:
            return self.current_user_id, True
        ok = self.select_user_workflow()
        return (self.current_user_id if ok else None), ok

    def select_user_workflow(self) -> bool:
        try:
            dlg = UserSelectDialog(self)
            result = dlg.exec_()
            if result != QDialog.Accepted:
                return False
            uid, uname = dlg.selected_user()
            if uid is None:
                return False
            self.current_user_id = uid
            self.current_user_name = uname
            self.show_all_forms()
            return True
        except Exception as e:
            QMessageBox.critical(self, "خطا", f"خطا در انتخاب کاربر:\n{str(e)}")
            return False

    def query_users_by_name(self, name_part: str):
        cursor = self.conn.cursor()
        norm = normalize_persian_text(name_part)

        try:
            sql = (
                "SELECT TOP 50 A.id AS Id, A.UserName AS Name "
                "FROM dbo.Authorize A WHERE "
                f"{sql_normalize_expr('A.UserName')} LIKE ? ORDER BY A.UserName;"
            )
            cursor.execute(sql, (build_like_param(norm),))
            rows = cursor.fetchall()
            if rows:
                return [(r[0], "" if r[1] is None else str(r[1])) for r in rows]

            for coll in candidate_collations():
                sql = (
                    "SELECT TOP 50 A.id AS Id, A.UserName AS Name "
                    "FROM dbo.Authorize A WHERE "
                    f"{sql_normalize_expr('A.UserName')} COLLATE {coll} LIKE ? "
                    "ORDER BY A.UserName;"
                )
                cursor.execute(sql, (build_like_param(norm),))
                rows = cursor.fetchall()
                if rows:
                    return [(r[0], "" if r[1] is None else str(r[1])) for r in rows]
        except Exception:
            pass

        table_candidates = ["Authorize", "Users", "User", "tblUsers", "tblUser"]
        id_candidates = ["ID", "Id", "UserId", "id"]
        name_candidates = ["UserName", "Username", "Name", "FullName"]
        for t in table_candidates:
            for idc in id_candidates:
                for nc in name_candidates:
                    try:
                        sql = (
                            f"SELECT TOP 50 {idc} AS Id, {nc} AS Name FROM dbo.{t} "
                            f"WHERE {sql_normalize_expr(nc)} LIKE ? ORDER BY {nc};"
                        )
                        cursor.execute(sql, (build_like_param(norm),))
                        rows = cursor.fetchall()
                        if rows:
                            return [(r[0], "" if r[1] is None else str(r[1])) for r in rows]

                        for coll in candidate_collations():
                            sql = (
                                f"SELECT TOP 50 {idc} AS Id, {nc} AS Name FROM dbo.{t} "
                                f"WHERE {sql_normalize_expr(nc)} COLLATE {coll} LIKE ? ORDER BY {nc};"
                            )
                            cursor.execute(sql, (build_like_param(norm),))
                            rows = cursor.fetchall()
                            if rows:
                                return [(r[0], "" if r[1] is None else str(r[1])) for r in rows]
                    except Exception:
                        continue
        return []

    def query_users_initial(self, limit: int = 200):
        cursor = self.conn.cursor()
        try:
            sql = (
                f"SELECT TOP {limit} A.id AS Id, A.UserName AS Name "
                "FROM dbo.Authorize A ORDER BY A.UserName;"
            )
            cursor.execute(sql)
            rows = cursor.fetchall()
            if rows:
                return [(r[0], "" if r[1] is None else str(r[1])) for r in rows]
        except Exception:
            pass

        table_candidates = ["Authorize", "Users", "User", "tblUsers", "tblUser"]
        id_candidates = ["ID", "Id", "UserId", "id"]
        name_candidates = ["UserName", "Username", "Name", "FullName"]
        for t in table_candidates:
            for idc in id_candidates:
                for nc in name_candidates:
                    try:
                        sql = (
                            f"SELECT TOP {limit} {idc} AS Id, {nc} AS Name FROM dbo.{t} "
                            f"ORDER BY {nc};"
                        )
                        cursor.execute(sql)
                        rows = cursor.fetchall()
                        if rows:
                            return [(r[0], "" if r[1] is None else str(r[1])) for r in rows]
                    except Exception:
                        continue
        return []

def start_application():
    app = QApplication(sys.argv)
    apply_app_icon(app)
    app.setStyleSheet(dark_theme)
    login = LoginWindow()
    login.show()
    return app.exec_()

if __name__ == "__main__":
    sys.exit(start_application())
